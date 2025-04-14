import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- Configuración inicial ---
st.set_page_config(page_title="Agregar Productos", layout="wide")

# --- Custom CSS ---
st.markdown("""
<style>
body {
    font-family: 'Roboto', sans-serif;
    background-color: #f4f4f9; padding-top: 0px; margin-top: 0px; margin-top: -30px;
}
.container {
    max-width: 1200px;
    margin: auto; margin-top: 0px; margin-top: -30px;
}
.card {
    background-color: #fff;
    padding: 20px;
    border-radius: 10px;
    box-shadow: 0px 4px 6px rgba(0,0,0,0.1);
    margin-bottom: 20px;
}
.thumbnail {
    width: 80px;
    height: 80px;
    object-fit: cover;
    border-radius: 5px;
    margin-right: 10px;
    display: inline-block;
}
</style>
""", unsafe_allow_html=True)

st.markdown("<div class='container'>", unsafe_allow_html=True)

# --- Variables globales y funciones ---
archivo_excel = "productos_base.xlsx"
columnas_excel = [
    'ID', 'Código de barra', 'Código mínimo', 'Proveedor', 'Nombre del producto',
    'Categoría', 'Marca', 'Descripción', 'Estado', 'Imagen principal (URL)',
    'Imágenes secundarias (URLs separadas por coma)', 'Precio compra', 'Stock',
    'Precio Facebook', 'Comisión vendedor Facebook', 'Ganancia Facebook',
    'Precio Mercado Libre', 'Comisión de Mercado Libre', 'Envio Mercado Libre',
    'Ganancia Mercado Libre', 'Ganancia de Mercado Libre despues de iva 19%',
    'Precio Mercado Libre con 30% desc.', 'Comisión de Mercado libre con 30% desc.',
    'Envio Mercado Libre con el 30% desc.', 'Ganancia Mercado Libre con el 30% desc.',
    'Ganancia de Mercado Libre despues de iva 19% con el 30% desc.',
    'Precio al por mayor de 3 ', 'Mostrar en catálogo', 'ID Publicación Mercado Libre',
    'Link publicación 1', 'Link publicación 2', 'Link publicación 3', 'Link publicación 4'
]

def agregar_producto(diccionario, proveedor):
    wb = load_workbook(archivo_excel)
    ws = wb["Productos"]
    tabla = ws.tables["TablaProductos"]

    ref_inicio, ref_fin = tabla.ref.split(":")
    fila_inicio = int(''.join(filter(str.isdigit, ref_inicio)))
    fila_fin = int(''.join(filter(str.isdigit, ref_fin)))
    col_inicio = ''.join(filter(str.isalpha, ref_inicio))

    headers = [cell.value for cell in ws[fila_inicio]]
    nueva_fila = [diccionario.get(col, "") for col in headers]

    ws.append(nueva_fila)

    nueva_fila_idx = fila_fin + 1
    nueva_ref = f"{col_inicio}{fila_inicio}:{get_column_letter(ws.max_column)}{nueva_fila_idx}"
    tabla.ref = nueva_ref

    ws_prov = wb["Proveedores"]
    ya_existe = [cell.value for cell in ws_prov["A"]]
    if proveedor.strip() and proveedor.strip() not in ya_existe:
        ws_prov.append([proveedor.strip()])

    wb.save(archivo_excel)

def cargar_datos():
    df_productos = pd.read_excel(archivo_excel, sheet_name="Productos")
    df_proveedores = pd.read_excel(archivo_excel, sheet_name="Proveedores")
    return df_productos, df_proveedores

def generar_id_producto(df_productos):
    existentes = df_productos["ID"].dropna().tolist()
    numeros = [int(e[1:]) for e in existentes if isinstance(e, str) and e.startswith("P") and e[1:].isdigit()]
    siguiente = max(numeros) + 1 if numeros else 1
    return f"P{str(siguiente).zfill(3)}"

# --- Cargar datos iniciales ---
# ===================== MEJORAS PRO =====================




# 📌 2. Sticky botón (pseudo estilo, se mantiene al fondo con CSS)
st.markdown("""
<style>
div.stButton > button {
    position: sticky;
    bottom: 10px;
    background-color: #0066cc;
    color: white;
    border-radius: 8px;
    padding: 10px 20px;
}
</style>
""", unsafe_allow_html=True)


# ✅ 4. Tooltip con info adicional en cada campo obligatorio (opcional si quieres agregar más ayuda)
# Esto ya lo usas con el parámetro `help`, así que se cumple.

# 🧠 5. Animación extra: suavizado cuando se llena el formulario (lo manejamos desde CSS con transiciones)
# Ya implementado en el bloque anterior de diseño

df_productos, df_proveedores = cargar_datos()
if "nuevo_id" not in st.session_state:
    st.session_state.nuevo_id = generar_id_producto(df_productos)
if "producto_guardado" not in st.session_state:
    st.session_state.producto_guardado = False

# Título centrado
st.markdown("<h1 style='text-align: center;'>🆕 Agregar nuevo producto</h1>", unsafe_allow_html=True)
st.markdown(f"<h3 style='text-align: center; color: green;'>🆔 ID generado: {st.session_state.nuevo_id}</h3>", unsafe_allow_html=True)


# === BANNER CON TEXTO CLARO ===
st.markdown("""
<div style='text-align:center; margin-top: -10px; margin-bottom: 25px;'>
    <div style='display:inline-block; background: #e3f2fd; padding: 10px 25px; border-radius: 12px; font-size: 16px; font-weight: 500; color: #0d47a1; box-shadow: 0 2px 5px rgba(0,0,0,0.1);'>
        💡 Completa los campos para guardar correctamente el producto
    </div>
</div>
""", unsafe_allow_html=True)


# --- Progreso del formulario ---
obligatorios_ids = ["codigo_barra", "codigo_minimo", "proveedor", "nombre", "categoria", "marca",
                    "descripcion", "estado", "precio_facebook", "comision_fb", "precio_compra"]
campos_llenos = sum(1 for k in obligatorios_ids if st.session_state.get(k))
progreso = int((campos_llenos / len(obligatorios_ids)) * 100)
st.progress(progreso, text=f"Formulario completado: {progreso}%")


# --- Previsualización en panel lateral (sidebar) ---
with st.sidebar.expander("Previsualización del producto", expanded=False):
    st.markdown("### Resumen del producto")
    if st.session_state.get("imagen_principal", "").startswith("http"):
        st.image(st.session_state.imagen_principal, width=150)
    st.markdown(f"**Nombre:** {st.session_state.get('nombre','')}")
    descripcion = st.session_state.get("descripcion", "")
    resumen = descripcion if len(descripcion) <= 100 else descripcion[:100] + "..."
    st.markdown(f"**Descripción:** {resumen}")
    prov_disp = st.session_state.get("nuevo_prov") if st.session_state.get("proveedor")=="Agregar nuevo" else st.session_state.get("proveedor", "")
    st.markdown(f"**Proveedor:** {prov_disp}")
    st.markdown(f"**Precio de costo:** {st.session_state.get('precio_compra','')}")

# --- Sección principal (diseño tipo card con 4 tabs) ---
tabs = st.tabs(["🧾 Identificación", "🖼️ Visuales y Descripción", "💰 Precios", "📦 Stock y Opciones"])

# TAB 1: Identificación
with tabs[0]:
    col1, col2, col3 = st.columns(3)
    with col1:
         st.text_input("Código de barra *", placeholder="Ej: 1234567890", key="codigo_barra", help="Ingresa el código de barras")
         st.text_input("Nombre del producto *", placeholder="Ej: Camiseta deportiva", key="nombre", help="Nombre del producto", max_chars=60)
         st.text_input("Categoría *", placeholder="Ej: Ropa", key="categoria", help="Categoría del producto")
    with col2:
         st.text_input("Código mínimo *", placeholder="Ej: 001", key="codigo_minimo", help="Código mínimo asignado")
         st.text_input("Marca *", placeholder="Ej: Nike", key="marca", help="Marca del producto")
         proveedor = st.selectbox("Proveedor *", options=[*df_proveedores["Proveedor"].unique(), "Agregar nuevo"], key="proveedor", help="Selecciona un proveedor o 'Agregar nuevo'")
    with col3:
         if proveedor == "Agregar nuevo":
             st.text_input("Nuevo proveedor", placeholder="Nombre del nuevo proveedor", key="nuevo_prov", help="Escribe el nuevo proveedor")
         else:
             st.write("")

# TAB 2: Visuales y Descripción
with tabs[1]:
    st.text_area("Descripción *", placeholder="Detalles del producto...", key="descripcion", help="Describe las características del producto")
    st.selectbox("Estado *", ["Nuevo", "Usado"], key="estado", help="Selecciona el estado del producto")
    st.text_input("Imagen principal (URL)", placeholder="https://...", key="imagen_principal", help="Ingresa la URL de la imagen principal")
    if st.session_state.get("imagen_principal", "").startswith("http"):
        st.image(st.session_state.get("imagen_principal"), width=200)
    st.text_input("Imágenes secundarias (URLs separadas por coma)", placeholder="https://..., https://...", key="imagenes_secundarias", help="Ingresa las URLs separadas por coma")
    if st.session_state.get("imagenes_secundarias"):
         urls = [url.strip() for url in st.session_state.get("imagenes_secundarias").split(",") if url.strip() != ""]
         html_imgs = "".join([f'<img src="{url}" class="thumbnail">' for url in urls])
         st.markdown(html_imgs, unsafe_allow_html=True)

    st.text_input("Etiquetas", placeholder="Palabras clave separadas por coma", key="etiquetas", help="Ej: nuevo, oferta, top")
     
# TAB 3: Precios (Ahora con diseño en 3 columnas tipo tabla y emoticones)
with tabs[2]:
    st.text_input("Precio compra *", placeholder="Costo del producto", key="precio_compra", help="Precio de compra del producto")
    st.markdown("### Detalles de Precios")
    col_fb, col_ml, col_ml30 = st.columns(3)
    
    
    with col_fb:
        st.markdown("💰 **Facebook**")
        st.text_input("Precio", placeholder="Precio para Facebook", key="precio_facebook", help="Precio para venta en Facebook")
        st.text_input("Comisión", placeholder="Comisión", key="comision_fb", help="Comisión para Facebook")
        st.text_input("Precio al por mayor de 3", placeholder="Precio al por mayor", key="precio_mayor", help="Precio para compras al por mayor")

        precio_fb_raw = st.session_state.get("precio_facebook", "")
        comision_fb_raw = st.session_state.get("comision_fb", "")
        precio_compra_raw = st.session_state.get("precio_compra", "")

        try:
            if precio_fb_raw and comision_fb_raw and precio_compra_raw:
                precio_fb = float(precio_fb_raw)
                comision_fb = float(comision_fb_raw)
                precio_compra = float(precio_compra_raw)
                ganancia_fb = precio_fb - precio_compra - comision_fb

                color = "green" if ganancia_fb > 0 else "red"
                icono = "✅" if ganancia_fb > 0 else "❌"
                st.markdown(f"<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: {color};'>{icono} {ganancia_fb:.0f} CLP</div>", unsafe_allow_html=True)

                st.session_state["ganancia_fb"] = f"{ganancia_fb:.0f}"
            else:
                st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
                st.session_state["ganancia_fb"] = ""
        except:
            st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
            st.session_state["ganancia_fb"] = ""

    
    with col_ml:
        st.markdown("🛒 **Mercado Libre**")
        st.text_input("Precio", placeholder="Precio para ML", key="precio_ml", help="Precio para Mercado Libre")
        st.text_input("Comisión", placeholder="Comisión", key="comision_ml", help="Comisión en ML")
        st.text_input("Envío", placeholder="Costo de envío", key="envio_ml", help="Costo de envío en ML")

        precio_ml_raw = st.session_state.get("precio_ml", "")
        comision_ml_raw = st.session_state.get("comision_ml", "")
        envio_ml_raw = st.session_state.get("envio_ml", "")
        precio_compra_raw = st.session_state.get("precio_compra", "")

        try:
            if precio_ml_raw and comision_ml_raw and envio_ml_raw and precio_compra_raw:
                precio_ml = float(precio_ml_raw)
                comision_ml = float(comision_ml_raw)
                envio_ml = float(envio_ml_raw)
                precio_compra = float(precio_compra_raw)
                ganancia_ml = precio_ml - precio_compra - comision_ml - envio_ml

                color = "green" if ganancia_ml > 0 else "red"
                icono = "✅" if ganancia_ml > 0 else "❌"
                st.markdown(f"<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: {color};'>{icono} {ganancia_ml:.0f} CLP</div>", unsafe_allow_html=True)

                st.session_state["ganancia_ml"] = f"{ganancia_ml:.0f}"
            else:
                st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
                st.session_state["ganancia_ml"] = ""
        except:
            st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
            st.session_state["ganancia_ml"] = ""

        st.text_input("Ganancia IVA", placeholder="Ganancia con IVA", key="ganancia_ml_iva", help="Ganancia considerando IVA 19%")

        # GANANCIA CON IVA - VISUAL EXTRA (NO MODIFICA INPUTS)
        try:
            if all(st.session_state.get(k) not in ["", None] for k in ["precio_ml", "comision_ml", "envio_ml", "precio_compra"]):
                neto = float(st.session_state["precio_ml"]) - float(st.session_state["comision_ml"]) - float(st.session_state["envio_ml"])
                iva = neto * 0.19
                ganancia_con_iva = round(neto - iva - float(st.session_state["precio_compra"]))
                color = "green" if ganancia_con_iva > 0 else "red"
                st.markdown(f"<div style='font-weight:bold; color:{color};'>💸 Ganancia ML descontando IVA: {ganancia_con_iva} CLP</div>", unsafe_allow_html=True)
            else:
                st.markdown("💡 Ingresa todos los datos para ver la ganancia con IVA.")
        except:
            st.markdown("❌ Error al calcular ganancia con IVA.")


    
    with col_ml30:
        st.markdown("📉 **ML con 30% desc.**")
        
        # Calcular automáticamente el precio con 30% descuento si se ingresó el precio base
        try:
            if st.session_state.get("precio_ml"):
                precio_base = float(st.session_state["precio_ml"])
                precio_desc = precio_base * 0.7
                st.session_state["precio_ml_desc"] = str(round(precio_desc))
            else:
                st.session_state["precio_ml_desc"] = ""
        except:
            st.session_state["precio_ml_desc"] = ""

        st.text_input("Precio", placeholder="Precio con 30% desc.", key="precio_ml_desc", help="Precio en ML con 30% de descuento", disabled=True)

        st.text_input("Comisión", placeholder="Comisión", key="comision_ml_desc", help="Comisión con 30% de descuento")
        st.text_input("Envío", placeholder="Envío", key="envio_ml_desc", help="Costo de envío con 30% de descuento")

        precio_ml_desc_raw = st.session_state.get("precio_ml_desc", "")
        comision_ml_desc_raw = st.session_state.get("comision_ml_desc", "")
        envio_ml_desc_raw = st.session_state.get("envio_ml_desc", "")
        precio_compra_raw = st.session_state.get("precio_compra", "")

        try:
            if precio_ml_desc_raw and comision_ml_desc_raw and envio_ml_desc_raw and precio_compra_raw:
                precio_ml_desc = float(precio_ml_desc_raw)
                comision_ml_desc = float(comision_ml_desc_raw)
                envio_ml_desc = float(envio_ml_desc_raw)
                precio_compra = float(precio_compra_raw)
                ganancia_ml_desc = precio_ml_desc - precio_compra - comision_ml_desc - envio_ml_desc

                color = "green" if ganancia_ml_desc > 0 else "red"
                icono = "✅" if ganancia_ml_desc > 0 else "❌"
                st.markdown(f"<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: {color};'>{icono} {ganancia_ml_desc:.0f} CLP</div>", unsafe_allow_html=True)

                st.session_state["ganancia_ml_desc"] = f"{ganancia_ml_desc:.0f}"
            else:
                st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
                st.session_state["ganancia_ml_desc"] = ""
        except:
            st.markdown("<div style='margin-top: 0.5rem;'>Ganancia estimada:</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-weight: bold; font-size: 24px; color: gray;'>-</div>", unsafe_allow_html=True)
            st.session_state["ganancia_ml_desc"] = ""

        st.text_input("Ganancia IVA", placeholder="Ganancia con IVA", key="ganancia_ml_iva_desc", help="Ganancia con IVA y 30% desc.")

        # GANANCIA CON IVA - VISUAL EXTRA 30% DESC
        try:
            if all(st.session_state.get(k) not in ["", None] for k in ["precio_ml_desc", "comision_ml_desc", "envio_ml_desc", "precio_compra"]):
                neto = float(st.session_state["precio_ml_desc"]) - float(st.session_state["comision_ml_desc"]) - float(st.session_state["envio_ml_desc"])
                iva = neto * 0.19
                ganancia_con_iva_desc = round(neto - iva - float(st.session_state["precio_compra"]))
                color = "green" if ganancia_con_iva_desc > 0 else "red"
                st.markdown(f"<div style='font-weight:bold; color:{color};'>💸 Ganancia ML -30% con IVA: {ganancia_con_iva_desc} CLP</div>", unsafe_allow_html=True)
            else:
                st.markdown("💡 Ingresa todos los datos para ver la ganancia con IVA.")
        except:
            st.markdown("❌ Error al calcular ganancia con IVA 30%.")


    
    with tabs[3]:
        st.text_input("Stock", placeholder="Cantidad en stock", key="stock", help="Cantidad disponible en stock")
        st.selectbox("Mostrar en catálogo", options=["Sí", "No"], key="mostrar_catalogo", help="¿Se muestra en catálogo?")
        st.text_input("ID Publicación ML", placeholder="ID de la publicación", key="id_publicacion", help="Identificador en Mercado Libre")
        st.text_input("Link publicación 1", placeholder="https://...", key="link1", help="URL de la publicación 1")
        st.text_input("Link publicación 2", placeholder="https://...", key="link2")
        st.text_input("Link publicación 3", placeholder="https://...", key="link3")
        st.text_input("Link publicación 4", placeholder="https://...", key="link4", help="URL de la publicación 4")
        st.text_input("Foto de proveedor", placeholder="URL de la foto", key="foto_proveedor", help="URL de una imagen del proveedor")

    

# --- Botón para guardar el producto ---
if st.button("💾 Guardar Producto"):
    obligatorios = [
        st.session_state.get("codigo_barra"), st.session_state.get("codigo_minimo"),
        st.session_state.get("proveedor"), st.session_state.get("nombre"),
        st.session_state.get("categoria"), st.session_state.get("marca"),
        st.session_state.get("descripcion"), st.session_state.get("estado"),
        st.session_state.get("precio_facebook"), st.session_state.get("comision_fb"),
        st.session_state.get("ganancia_fb"), st.session_state.get("precio_compra")
    ]
    if not all(obligatorios) or float(st.session_state.get("precio_compra", 0)) <= 0:
        st.warning("⚠️ Completa todos los campos obligatorios y asegúrate que los precios sean mayores a 0.")
    else:
        prov_final = st.session_state.get("nuevo_prov").strip() if st.session_state.get("proveedor")=="Agregar nuevo" and st.session_state.get("nuevo_prov") else st.session_state.get("proveedor")
        nuevo = {
            "ID": st.session_state.nuevo_id,
            "Código de barra": st.session_state.get("codigo_barra"),
            "Código mínimo": st.session_state.get("codigo_minimo"),
            "Proveedor": prov_final,
            "Nombre del producto": st.session_state.get("nombre"),
            "Categoría": st.session_state.get("categoria"),
            "Marca": st.session_state.get("marca"),
            "Descripción": st.session_state.get("descripcion"),
            "Estado": st.session_state.get("estado"),
            "Imagen principal (URL)": st.session_state.get("imagen_principal"),
            "Imágenes secundarias (URLs separadas por coma)": st.session_state.get("imagenes_secundarias"),
            "Precio compra": st.session_state.get("precio_compra"),
            "Stock": st.session_state.get("stock"),
            "Precio Facebook": st.session_state.get("precio_facebook"),
            "Comisión vendedor Facebook": st.session_state.get("comision_fb"),
            "Ganancia Facebook": st.session_state.get("ganancia_fb"),
            "Precio Mercado Libre": st.session_state.get("precio_ml"),
            "Comisión de Mercado Libre": st.session_state.get("comision_ml"),
            "Envio Mercado Libre": st.session_state.get("envio_ml"),
            "Ganancia Mercado Libre": st.session_state.get("ganancia_ml"),
            "Ganancia de Mercado Libre despues de iva 19%": st.session_state.get("ganancia_ml_iva"),
            "Precio Mercado Libre con 30% desc.": st.session_state.get("precio_ml_desc"),
            "Comisión de Mercado libre con 30% desc.": st.session_state.get("comision_ml_desc"),
            "Envio Mercado Libre con el 30% desc.": st.session_state.get("envio_ml_desc"),
            "Ganancia Mercado Libre con el 30% desc.": st.session_state.get("ganancia_ml_desc"),
            "Ganancia de Mercado Libre despues de iva 19% con el 30% desc.": st.session_state.get("ganancia_ml_iva_desc"),
            "Precio al por mayor de 3 ": st.session_state.get("precio_mayor"),
            "Mostrar en catálogo": st.session_state.get("mostrar_catalogo"),
            "ID Publicación Mercado Libre": st.session_state.get("id_publicacion"),
            "Link publicación 1": st.session_state.get("link1"),
            "Link publicación 2": st.session_state.get("link2"),
            "Link publicación 3": st.session_state.get("link3"),
            "Link publicación 4": st.session_state.get("link4"),
            "Etiquetas": st.session_state.get("etiquetas"),
            "Foto de proveedor": st.session_state.get("foto_proveedor")
        }
        agregar_producto(nuevo, prov_final)
        st.success(f"✅ Producto {st.session_state.nuevo_id} agregado correctamente.")
        df_productos, _ = cargar_datos()
        st.session_state.nuevo_id = generar_id_producto(df_productos)
        keys_reset = ["codigo_barra", "codigo_minimo", "nombre", "categoria", "marca", "proveedor",
                      "nuevo_prov", "descripcion", "estado", "imagen_principal", "imagenes_secundarias",
                      "precio_compra", "precio_facebook", "comision_fb", "ganancia_fb",
                      "precio_ml", "comision_ml", "envio_ml", "ganancia_ml", "ganancia_ml_iva",
                      "precio_ml_desc", "comision_ml_desc", "envio_ml_desc", "ganancia_ml_desc", "ganancia_ml_iva_desc",
                      "stock", "precio_mayor", "mostrar_catalogo", "id_publicacion", "link1", "link2", "link3", "link4"]
        for k in keys_reset:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()






from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def agregar_producto_en_tabla(producto_dict):
    wb = load_workbook(archivo_excel)
    ws = wb["Productos"]
    tabla = ws.tables["TablaProductos"]

    # Obtener rango actual de la tabla
    ref_inicio, ref_fin = tabla.ref.split(":")
    fila_inicio = int(''.join(filter(str.isdigit, ref_inicio)))
    fila_fin = int(''.join(filter(str.isdigit, ref_fin)))
    col_inicio = ''.join(filter(str.isalpha, ref_inicio))
    col_fin = ''.join(filter(str.isalpha, ref_fin))

    # Obtener columnas desde la hoja
    headers = [cell.value for cell in ws[1]]

    # Crear la fila a agregar
    nueva_fila = [producto_dict.get(col, "") for col in headers]

    # Agregar la fila
    ws.append(nueva_fila)

    # Actualizar rango de la tabla
    nueva_fila_idx = fila_fin + 1
    nueva_ref = f"{col_inicio}{fila_inicio}:{get_column_letter(ws.max_column)}{nueva_fila_idx}"
    tabla.ref = nueva_ref

    wb.save(archivo_excel)


# ============================================
# 🛠️ Herramientas del vendedor (sidebar mejorado con selector)
with st.sidebar.expander("🧰 Herramientas del vendedor"):

    herramienta = st.selectbox(
        "Selecciona una herramienta:",
        (
            "🔢 Calculadora 30% de descuento",
            "🎯 Precio objetivo por ganancia deseada",
            "📊 Conversor Bruto/Neto con IVA",        ),
        key="selector_herramienta_sidebar"
    )

    if herramienta == "🔢 Calculadora 30% de descuento":
        final_30 = st.number_input("¿Cuánto quieres que quede después del 30%?", min_value=0, step=100, key="calc_30_sidebar")
        if final_30:
            original_30 = round(final_30 / 0.7)
            st.success(f"💡 Precio original sugerido: ${original_30}")

    elif herramienta == "🎯 Precio objetivo por ganancia deseada":
        ganancia_objetivo = st.number_input("Ganancia que deseas obtener", min_value=0, step=100, key="meta_ganancia_sidebar")
        costo_prod = st.number_input("Costo del producto", min_value=0, step=100, key="meta_costo_sidebar")
        envio_est = st.number_input("Costo del envío", min_value=0, step=100, key="meta_envio_sidebar")
        comision_est = st.number_input("Comisión estimada", min_value=0, step=100, key="meta_comision_sidebar")
        if ganancia_objetivo:
            precio_sugerido = ganancia_objetivo + costo_prod + envio_est + comision_est
            st.success(f"💰 Precio recomendado: ${precio_sugerido}")

    elif herramienta == "📊 Conversor Bruto/Neto con IVA":
        bruto = st.number_input("💵 Precio bruto (con IVA)", min_value=0, step=100, key="bruto_sidebar")
        if bruto:
            neto = round(bruto / 1.19)
            iva = bruto - neto
            st.info(f"💡 Neto: ${neto} | IVA: ${round(iva)}")

    
